import pandas as pd
import pdfplumber
import re
import os
import csv
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def read_keywords_from_csv(csv_path):
    """Read keywords from a CSV file"""
    keywords = []
    try:
        with open(csv_path, 'r', encoding='cp1252') as file:
            csv_reader = csv.reader(file)
            for row in csv_reader:
                if row:
                    keywords.append(row[0].strip())
        print(f"Loaded {len(keywords)} keywords from {csv_path}")
        return keywords
    except Exception as e:
        print(f"Error reading CSV file: {e}")
        return []

def extract_keywords_from_pdf(pdf_path, keywords):
    """Find keywords in PDF and create a mapping of keywords to page numbers"""
    keyword_pages = {}
    
    print(f"Processing PDF: {pdf_path}")
    print(f"Searching for keywords...")
    
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            page_num = i + 1
            try:
                text = page.extract_text()
                if text:
                    for keyword in keywords:
                        if keyword.lower() in text.lower():
                            if keyword not in keyword_pages:
                                keyword_pages[keyword] = []
                            keyword_pages[keyword].append(page_num)
                            print(f"Found keyword '{keyword}' on page {page_num}")
            except Exception as e:
                print(f"Error on page {page_num}: {e}")
    
    return keyword_pages

def create_standardized_data(keyword_pages, company_name="Philips"):
    """Create standardized data from keyword to page mapping"""
    data = []
    
    for keyword, pages in keyword_pages.items():
        entry = {
            'name': company_name,
            'category': 'DR',
            'variable': keyword,
            'value': 1,
        }
        
        # Add page references (up to 5)
        for i, page in enumerate(pages[:5]):
            entry[f'Page_ref{i+1}'] = page
        
        data.append(entry)
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Ensure all Page_ref columns exist (up to 5)
    for i in range(1, 6):
        col_name = f'Page_ref{i}'
        if col_name not in df.columns:
            df[col_name] = ''
    
    return df

def create_formatted_excel(df, output_path):
    """Create a formatted Excel file in the style shown in the example"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DR Data"
    
    # Add title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "How my ideal table looks like..."
    title_cell.font = Font(name='Arial', size=24, bold=True, color="9C0006")
    title_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Set row height for title
    ws.row_dimensions[1].height = 50
    
    # Define header row
    headers = ['name', 'category', 'variable', 'value', 'Page_ref1', 'Page_ref2', 'Page_ref3', 'Page_ref4', 'Page_ref5']
    
    # Add header row
    header_row = 3
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # Add borders to header
    thin_border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='thin'))
    
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=header_row, column=col_idx).border = thin_border
    
    # Add data rows
    for row_idx, (_, row_data) in enumerate(df.iterrows(), 1):
        for col_idx, header in enumerate(headers, 1):
            if header in row_data:
                ws.cell(row=header_row + row_idx, column=col_idx).value = row_data[header]
                ws.cell(row=header_row + row_idx, column=col_idx).border = thin_border
    
    # Add "Additional data items" section
    additional_row = header_row
    ws.cell(row=additional_row, column=len(headers) + 2).value = "Additional data items"
    ws.cell(row=additional_row, column=len(headers) + 2).font = Font(name='Arial', size=20, bold=True)
    
    # Add bullet points for additional items
    additional_items = [
        "Assurance level",
        "Results of materiality assessment (DMA)",
        "Location within report",
        "Any other additional items you can standardize",
        "Carbon Offset E1-7"
    ]
    
    for idx, item in enumerate(additional_items):
        bullet_row = additional_row + idx + 2
        ws.cell(row=bullet_row, column=len(headers) + 2).value = "• " + item
        ws.cell(row=bullet_row, column=len(headers) + 2).font = Font(name='Arial', size=14)
    
    # Auto-adjust column widths
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15
    
    # Adjust width for additional items column
    ws.column_dimensions[get_column_letter(len(headers) + 2)].width = 35
    
    # Save the workbook
    wb.save(output_path)
    print(f"Formatted Excel file created at: {output_path}")

def main():
    # File paths
    csv_path = 'DR_list.csv'
    pdf_directory = 'pdfs'
    output_directory = 'output'
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_directory):
        os.makedirs(output_directory)
    
    # Read keywords from CSV file
    keywords = read_keywords_from_csv(csv_path)
    if not keywords:
        print("No keywords found or error reading CSV file.")
        return
    
    # Process each PDF file in the directory
    for pdf_file in os.listdir(pdf_directory):
        if pdf_file.lower().endswith('.pdf'):
            pdf_path = os.path.join(pdf_directory, pdf_file)
            base_name = os.path.splitext(pdf_file)[0]
            
            # Extract keywords and find their pages
            keyword_pages = extract_keywords_from_pdf(pdf_path, keywords)
            
            if keyword_pages:
                # Create standardized data
                standardized_data = create_standardized_data(keyword_pages)
                
                # Create intermediate file (optional)
                intermediate_excel = os.path.join(output_directory, f"{base_name}_tables_raw.xlsx")
                standardized_data.to_excel(intermediate_excel, index=False)
                print(f"Raw data saved to {intermediate_excel}")
                
                # Create formatted Excel
                output_excel = os.path.join(output_directory, f"{base_name}_tables_formatted.xlsx")
                create_formatted_excel(standardized_data, output_excel)
            else:
                print(f"No keywords found in {pdf_file}.")

if __name__ == "__main__":
    main()